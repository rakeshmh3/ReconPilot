import os
import glob
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return None
        # Try parsing various common formats
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
            try:
                return datetime.datetime.strptime(val_str, fmt)
            except ValueError:
                continue
    return None

def main():
    recon_folder = 'Recon_Files'
    pattern = os.path.join(recon_folder, '*.xlsx')
    files = glob.glob(pattern)
    
    records = []
    
    for f in sorted(files):
        filename = os.path.basename(f)
        p_by, p_dt, r_by, r_dt = None, None, None, None
        compliance1, compliance2 = None, None
        remarks = []
        
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            if 'SignOff' not in wb.sheetnames:
                remarks.append("Missing SignOff worksheet")
                compliance1 = False
                compliance2 = False
            else:
                ws = wb['SignOff']
                p_by = ws['D30'].value
                p_dt = ws['D31'].value
                r_by = ws['G30'].value
                r_dt = ws['G31'].value
                
                # Check Performed By / Reviewed By compliance
                p_by_str = str(p_by).strip() if p_by is not None else ""
                r_by_str = str(r_by).strip() if r_by is not None else ""
                
                if not p_by_str or not r_by_str:
                    compliance1 = False
                    remarks.append("Performed_By or Reviewed_By is empty")
                else:
                    compliance1 = (p_by_str.lower() != r_by_str.lower())
                
                # Check Dates compliance
                p_dt_parsed = parse_date(p_dt)
                r_dt_parsed = parse_date(r_dt)
                
                if p_dt is not None and p_dt_parsed is None:
                    remarks.append(f"Failed to parse Performed_Date: {p_dt}")
                if r_dt is not None and r_dt_parsed is None:
                    remarks.append(f"Failed to parse Reviewed_Date: {r_dt}")
                
                if p_dt_parsed is None or r_dt_parsed is None:
                    compliance2 = False
                    if p_dt is None or r_dt is None:
                        remarks.append("Performed_Date or Reviewed_Date is empty")
                else:
                    compliance2 = (r_dt_parsed >= p_dt_parsed)
                    
        except Exception as e:
            remarks.append(f"Error reading workbook: {str(e)}")
            compliance1 = False
            compliance2 = False
            
        remarks_str = "; ".join(remarks) if remarks else None
        
        records.append({
            'Workbook_Name': filename,
            'Performed_By': p_by,
            'Performed_Date': p_dt,
            'Reviewed_By': r_by,
            'Reviewed_Date': r_dt,
            'compliance1': compliance1,
            'compliance2': compliance2,
            'Remarks': remarks_str
        })
        
    df = pd.DataFrame(records)
    
    os.makedirs('reports', exist_ok=True)
    report_path = os.path.join('reports', 'Recon_Compliance_Report.xlsx')
    
    df.to_excel(report_path, index=False)
    
    # Reload and style
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    
    # Find columns index of compliance1 and compliance2
    headers = [cell.value for cell in ws[1]]
    col_c1_idx = headers.index('compliance1') + 1
    col_c2_idx = headers.index('compliance2') + 1
    
    for row in range(2, ws.max_row + 1):
        cell_c1 = ws.cell(row=row, column=col_c1_idx)
        if cell_c1.value is False or str(cell_c1.value).strip().upper() == "FALSE":
            cell_c1.fill = red_fill
            cell_c1.font = red_font
            
        cell_c2 = ws.cell(row=row, column=col_c2_idx)
        if cell_c2.value is False or str(cell_c2.value).strip().upper() == "FALSE":
            cell_c2.fill = red_fill
            cell_c2.font = red_font
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(report_path)
    print(f"Compliance report generated successfully at: {report_path}")

if __name__ == '__main__':
    main()
